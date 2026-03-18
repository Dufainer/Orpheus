#!/usr/bin/env python3
"""Orpheus — Music Metadata Updater
Updates metadata, lyrics and cover art for local FLAC/MP3 files
and generates an Excel report at the end.

Requires:
    pip install mutagen requests openpyxl pillow PyQt5

APIs used (all free, no key needed):
    - MusicBrainz  → metadata (title, artist, genre, year)
    - LRCLIB       → lyrics
    - Cover Art Archive → cover art
"""

import os
import sys
import struct
import threading
from pathlib import Path
from datetime import datetime

import mutagen
from mutagen.flac import FLAC
from mutagen.id3 import ID3
from mutagen.mp3 import MP3
import openpyxl

from config import *
from utils import *
from metadata import *
from lyrics import *
from gui import run_gui

# ─── FLAC UTILITIES ───────────────────────────────────────────────────────────

def read_flac(filepath):
    """Reads blocks from a FLAC file."""
    blocks = []
    with open(filepath, 'rb') as f:
        if f.read(4) != b'fLaC':
            raise ValueError("Invalid FLAC file")
        while True:
            header = f.read(4)
            if len(header) < 4:
                break
            block_type = header[0] & 0x7F
            is_last = (header[0] & 0x80) != 0
            block_size = struct.unpack('>I', b'\x00' + header[1:4])[0]
            data = f.read(block_size)
            blocks.append({'type': block_type, 'data': data, 'is_last': is_last})
            if is_last:
                break
        audio_data = f.read()
    return blocks, audio_data

def write_flac(filepath, blocks, audio_data):
    """Writes blocks back to a FLAC file."""
    with open(filepath, 'wb') as f:
        f.write(b'fLaC')
        for i, block in enumerate(blocks):
            is_last = (i == len(blocks) - 1)
            type_byte = block['type'] | (0x80 if is_last else 0x00)
            size_bytes = struct.pack('>I', len(block['data']))[1:]
            f.write(bytes([type_byte]) + size_bytes)
            f.write(block['data'])
        f.write(audio_data)

def parse_vorbis(data):
    """Parses a Vorbis Comment block from a FLAC file."""
    pos = 0
    vendor_len = struct.unpack('<I', data[pos:pos+4])[0]
    pos += 4
    vendor = data[pos:pos+vendor_len].decode('utf-8', errors='replace')
    pos += vendor_len
    count = struct.unpack('<I', data[pos:pos+4])[0]
    pos += 4
    tags = {}
    for _ in range(count):
        length = struct.unpack('<I', data[pos:pos+4])[0]
        pos += 4
        comment = data[pos:pos+length].decode('utf-8', errors='replace')
        pos += length
        if '=' in comment:
            k, v = comment.split('=', 1)
            tags[k.upper()] = v
    return vendor, tags

def build_vorbis(vendor, tags):
    """Builds a Vorbis Comment block."""
    vendor_bytes = vendor.encode('utf-8')
    result = struct.pack('<I', len(vendor_bytes)) + vendor_bytes
    comments = []
    for k, v in tags.items():
        comment = f"{k}={v}"
        comment_bytes = comment.encode('utf-8')
        comments.append(struct.pack('<I', len(comment_bytes)) + comment_bytes)
    result += struct.pack('<I', len(comments))
    result += b''.join(comments)
    return result

# ─── TAG FUNCTIONS ────────────────────────────────────────────────────────────

def get_flac_tags(filepath):
    """Gets tags from a FLAC file."""
    try:
        audio = FLAC(filepath)
        tags = {}
        for key, value in audio.tags.items():
            if isinstance(value, list) and value:
                tags[key] = str(value[0])
            else:
                tags[key] = str(value)
        return tags
    except Exception:
        return {}

def save_flac_tags(filepath, tags):
    """Saves tags to a FLAC file."""
    try:
        audio = FLAC(filepath)
        for key, value in tags.items():
            audio[key] = value
        audio.save()
    except Exception as e:
        print(f"Error saving FLAC: {e}")

def get_mp3_tags(filepath):
    """Gets tags from an MP3 file."""
    try:
        audio = MP3(filepath, ID3=ID3)
        tags = {}
        for key, value in audio.tags.items():
            if hasattr(value, 'text'):
                tags[key] = str(value.text[0]) if value.text else ''
            else:
                tags[key] = str(value)
        return tags
    except Exception:
        return {}

def save_mp3_tags(filepath, tags):
    """Saves tags to an MP3 file."""
    try:
        audio = MP3(filepath, ID3=ID3)
        for key, value in tags.items():
            if key in ['TITLE', 'ARTIST', 'ALBUM', 'GENRE', 'DATE', 'TRACKNUMBER']:
                frame_class = getattr(ID3, f'T{key[:4]}') if key != 'TRACKNUMBER' else ID3.TRCK
                audio.tags.add(frame_class(encoding=3, text=value))
        audio.save()
    except Exception as e:
        print(f"Error saving MP3: {e}")

def has_cover_art(filepath):
    """Checks if a FLAC file has cover art."""
    if filepath.suffix.lower() == '.flac':
        try:
            audio = FLAC(filepath)
            for picture in audio.pictures:
                if picture.type == 3:  # Cover (front)
                    return True
        except:
            pass
    return False

def save_cover_art(filepath, cover_data):
    """Saves cover art to a FLAC file."""
    if filepath.suffix.lower() == '.flac' and cover_data:
        try:
            audio = FLAC(filepath)
            from mutagen.flac import Picture
            pic = Picture()
            pic.type = 3  # Cover (front)
            pic.mime = 'image/jpeg'
            pic.desc = 'Cover'
            pic.data = cover_data
            audio.add_picture(pic)
            audio.save()
        except Exception as e:
            print(f"Error saving cover art: {e}")

# ─── PROCESSING ───────────────────────────────────────────────────────────────

def process_file(filepath, print_func=print):
    """Processes a single music file."""
    ext = filepath.suffix.lower()

    result = {
        'file': filepath,
        'changes': [],
        'errors': []
    }

    print_func(f"\n  🎵 {filepath.name}")

    # Get current tags
    if ext == '.flac':
        tags = get_flac_tags(filepath)
        try:
            audio = FLAC(filepath)
            result['quality'] = f"{audio.info.bits_per_sample}-bit/{audio.info.sample_rate//1000:.0f}kHz"
        except:
            result['quality'] = 'FLAC'
        already_has_cover = has_cover_art(filepath)
    else:
        tags = get_mp3_tags(filepath)
        result['quality'] = 'MP3'
        already_has_cover = False

    artist = tags.get('ARTIST', '').strip()
    title = tags.get('TITLE', '').strip()
    album = tags.get('ALBUM', '').strip()

    result['artist'] = artist
    result['title'] = title
    result['album'] = album

    if not artist or not title:
        # Try to extract from filename
        name = filepath.stem.replace('_', ' ')
        if ' - ' in name:
            parts = name.split(' - ', 1)
            artist = artist or parts[0].strip()
            title = title or parts[1].strip()
        else:
            title = title or name

        if not artist:
            print_func(f"    ⚠️  No artist tag found — using parent folder name")
            artist = filepath.parent.parent.name

    print_func(f"    👤 {artist} — {title}")
    if album:
        print_func(f"    💿 {album}  [{result['quality']}]")

    new_tags = {}
    cover_data = None

    # 1. Search metadata on MusicBrainz
    if artist and title:
        print_func(f"    🔍 Searching MusicBrainz...")
        mb_data = search_musicbrainz(artist, title)

        needs_cover = not already_has_cover or OVERWRITE_COVER

        if mb_data:
            # Year
            if 'DATE' in mb_data:
                old_date = tags.get('DATE', '')
                if not old_date or OVERWRITE_DATE:
                    new_tags['DATE'] = mb_data['DATE']
                    label = f"Year: {old_date} → {mb_data['DATE']}" if old_date else f"Year: {mb_data['DATE']}"
                    result['changes'].append(label)
                    print_func(f"    ✅ 📅 {label}")

            # Genre
            if 'GENRE' in mb_data:
                old_genre = tags.get('GENRE', '')
                if not old_genre or OVERWRITE_GENRE:
                    new_tags['GENRE'] = mb_data['GENRE']
                    label = f"Genre: {old_genre} → {mb_data['GENRE']}" if old_genre else f"Genre: {mb_data['GENRE']}"
                    result['changes'].append(label)
                    print_func(f"    ✅ 🎸 {label}")

            # Album
            if 'ALBUM' in mb_data:
                old_album = tags.get('ALBUM', '')
                if not old_album or OVERWRITE_ALBUM:
                    new_tags['ALBUM'] = mb_data['ALBUM']
                    label = f"Album: {old_album} → {mb_data['ALBUM']}" if old_album else f"Album: {mb_data['ALBUM']}"
                    result['changes'].append(label)
                    print_func(f"    ✅ 💿 {label}")

            # Track number
            if 'TRACKNUMBER' in mb_data:
                old_track = tags.get('TRACKNUMBER', '')
                if not old_track or OVERWRITE_TRACK:
                    new_tags['TRACKNUMBER'] = mb_data['TRACKNUMBER']
                    label = f"Track: {old_track} → {mb_data['TRACKNUMBER']}" if old_track else f"Track: {mb_data['TRACKNUMBER']}"
                    result['changes'].append(label)
                    print_func(f"    ✅ 🔢 {label}")

            # Cover art
            if needs_cover and mb_data.get('_mbid'):
                print_func(f"    🖼️  Downloading cover art...")
                cover_data = fetch_cover_art(mb_data['_mbid'])
                if cover_data:
                    result['changes'].append("Cover art updated")
                    print_func(f"    ✅ 🖼️  Cover art downloaded ({len(cover_data)//1024} KB)")
                else:
                    print_func(f"    ⚠️  Cover art not found in Cover Art Archive")
        else:
            print_func(f"    ⚠️  Not found on MusicBrainz")

            # Fallback to iTunes
            print_func(f"    🔍 Searching iTunes...")
            itunes_data = search_itunes(artist, title)

            if itunes_data:
                # Year
                if 'DATE' in itunes_data:
                    old_date = tags.get('DATE', '')
                    if not old_date or OVERWRITE_DATE:
                        new_tags['DATE'] = itunes_data['DATE']
                        label = f"Year: {old_date} → {itunes_data['DATE']}" if old_date else f"Year: {itunes_data['DATE']}"
                        result['changes'].append(label)
                        print_func(f"    ✅ 📅 {label}")

                # Genre
                if 'GENRE' in itunes_data:
                    old_genre = tags.get('GENRE', '')
                    if not old_genre or OVERWRITE_GENRE:
                        new_tags['GENRE'] = itunes_data['GENRE']
                        label = f"Genre: {old_genre} → {itunes_data['GENRE']}" if old_genre else f"Genre: {itunes_data['GENRE']}"
                        result['changes'].append(label)
                        print_func(f"    ✅ 🎸 {label}")

                # Album
                if 'ALBUM' in itunes_data:
                    old_album = tags.get('ALBUM', '')
                    if not old_album or OVERWRITE_ALBUM:
                        new_tags['ALBUM'] = itunes_data['ALBUM']
                        label = f"Album: {old_album} → {itunes_data['ALBUM']}" if old_album else f"Album: {itunes_data['ALBUM']}"
                        result['changes'].append(label)
                        print_func(f"    ✅ 💿 {label}")

                # Track number
                if 'TRACKNUMBER' in itunes_data:
                    old_track = tags.get('TRACKNUMBER', '')
                    if not old_track or OVERWRITE_TRACK:
                        new_tags['TRACKNUMBER'] = itunes_data['TRACKNUMBER']
                        label = f"Track: {old_track} → {itunes_data['TRACKNUMBER']}" if old_track else f"Track: {itunes_data['TRACKNUMBER']}"
                        result['changes'].append(label)
                        print_func(f"    ✅ 🔢 {label}")

                # Cover art from iTunes
                if needs_cover and 'cover_url' in itunes_data:
                    print_func(f"    🖼️  Downloading cover art from iTunes...")
                    cover_data = fetch_bytes(itunes_data['cover_url'])
                    if cover_data:
                        result['changes'].append("Cover art updated")
                        print_func(f"    ✅ 🖼️  Cover art downloaded ({len(cover_data)//1024} KB)")
                    else:
                        print_func(f"    ⚠️  Cover art not found")
            else:
                print_func(f"    ⚠️  Not found on iTunes either")

    # 2. Search for lyrics
    has_lyrics = bool(tags.get('LYRICS') or tags.get('UNSYNCEDLYRICS'))
    if not has_lyrics or OVERWRITE_LYRICS:
        duration = 0
        print_func(f"    📝 Searching for lyrics on LRCLIB...")
        lyrics = fetch_lyrics_lrclib(artist, title, album, duration)

        if not lyrics:
            print_func(f"    📝 Trying lyrics.ovh...")
            lyrics = fetch_lyrics_fallback(artist, title)

        if not lyrics:
            print_func(f"    📝 Trying ChartLyrics...")
            lyrics = fetch_lyrics_chartlyrics(artist, title)

        if lyrics:
            new_tags['LYRICS'] = lyrics
            new_tags['UNSYNCEDLYRICS'] = lyrics
            label = "Lyrics updated" if has_lyrics else "Lyrics added"
            result['changes'].append(label)
            print_func(f"    ✅ 📝 {label} ({len(lyrics)} chars)")
        else:
            print_func(f"    ⚠️  Lyrics not found")

    # Save changes
    if new_tags or cover_data:
        if not DRY_RUN:
            if ext == '.flac':
                save_flac_tags(filepath, new_tags)
                if cover_data:
                    save_cover_art(filepath, cover_data)
            else:
                save_mp3_tags(filepath, new_tags)
            print_func(f"    💾 Saved successfully  ({len(new_tags)} tag(s) updated)")
        else:
            print_func(f"    🔒 Dry-run mode — file not modified")

    return result

# ─── EXCEL REPORT ─────────────────────────────────────────────────────────────

def generate_excel_report(results, report_path):
    """Generates an Excel report with the results."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Music Report"

    headers = ["File", "Artist", "Title", "Album", "Changes", "Errors", "Quality"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=str(result['file']))
        ws.cell(row=row, column=2, value=result.get('artist', ''))
        ws.cell(row=row, column=3, value=result.get('title', ''))
        ws.cell(row=row, column=4, value=result.get('album', ''))
        ws.cell(row=row, column=5, value='; '.join(result['changes']))
        ws.cell(row=row, column=6, value='; '.join(result['errors']))
        ws.cell(row=row, column=7, value=result.get('quality', ''))

    wb.save(report_path)
    print(f"Report saved: {report_path}")

_CHANGE_EMOJI = {
    'year': '📅', 'date': '📅',
    'genre': '🎸',
    'album': '💿',
    'track': '🔢',
    'cover': '🖼️',
    'lyric': '📝',
}

def _emoji_for_change(text):
    low = text.lower()
    for key, emoji in _CHANGE_EMOJI.items():
        if key in low:
            return emoji
    return '✅'

def print_change_summary(results):
    """Prints a detailed change summary to the console."""
    updated   = [r for r in results if r['changes']]
    no_change = [r for r in results if not r['changes'] and not r['errors']]
    errors    = [r for r in results if r['errors']]

    print(f"\n{'═' * 60}")
    print(f"  📊 PROCESSING SUMMARY")
    print(f"{'═' * 60}")
    print(f"  🎼 Total     : {len(results)}")
    print(f"  ✅ Updated   : {len(updated)}")
    print(f"  ➖ Unchanged : {len(no_change)}")
    print(f"  ❌ Errors    : {len(errors)}")
    print(f"{'═' * 60}")

    if updated:
        print(f"\n  ✅ UPDATED FILES ({len(updated)})")
        print(f"  {'─' * 56}")
        for r in updated:
            name = r['file'].name
            artist = r.get('artist', '')
            label = f"{artist} — {name}" if artist else name
            print(f"  🎵 {label}")
            for change in r['changes']:
                print(f"       {_emoji_for_change(change)} {change}")

    if errors:
        print(f"\n  ❌ ERRORS ({len(errors)})")
        print(f"  {'─' * 56}")
        for r in errors:
            print(f"  ❗ {r['file'].name}")
            for err in r['errors']:
                print(f"       {err}")

    if no_change:
        print(f"\n  ➖ NO CHANGES ({len(no_change)})")
        print(f"  {'─' * 56}")
        for r in no_change:
            print(f"  ◦  {r['file'].name}")

    print(f"\n{'═' * 60}")

# ─── MAIN FUNCTION ────────────────────────────────────────────────────────────

def main(music_dir=None, print_func=print, progress_callback=None,
         pause_event=None, song_card_callback=None, ask_report=False,
         stop_event=None):
    """
    Core processing function.
    When ask_report=True the user is prompted after processing whether to
    generate the Excel report (used by the console menu).
    """
    if music_dir is None:
        music_dir = Path(MUSIC_DIR)
    else:
        music_dir = Path(music_dir)

    if not music_dir.exists():
        print_func(f"❌ Folder not found: {music_dir}")
        if music_dir == Path(MUSIC_DIR):
            print_func("   Edit MUSIC_DIR in config.py")
        sys.exit(1)

    # Find all music files
    files = []
    for ext in EXTENSIONS:
        files.extend(music_dir.rglob(f"*{ext}"))
    files = sorted(files)

    print_func("🎵 Orpheus — Music Metadata Updater")
    print_func("═" * 50)
    print_func(f"📂 Folder : {music_dir}")
    print_func(f"🎼 Files  : {len(files)}")
    print_func("═" * 50)

    results = []
    ok = 0
    for i, filepath in enumerate(files, 1):
        # Check if processing should be stopped
        if stop_event and stop_event.is_set():
            print_func("⏹️  Processing stopped by user")
            break

        # Check if processing should be paused
        if pause_event and not pause_event.is_set():
            print_func(f"\n⏸️  Paused at file {i}/{len(files)}")
            pause_event.wait()
            # After unblocking, check if we were stopped (not just resumed)
            if stop_event and stop_event.is_set():
                print_func("⏹️  Processing stopped by user")
                break
            print_func(f"▶️  Resuming from file {i}/{len(files)}")

        print_func(f"\n─── [{i}/{len(files)}] ───────────────────────────────────")

        if song_card_callback:
            song_card_callback({
                'name': filepath.name,
                'status': 'processing',
                'artist': '', 'title': '', 'album': '',
                'changes': [], 'errors': [],
                'duration': '',
                'file_size': f"{filepath.stat().st_size / (1024*1024):.1f}MB"
            })

        try:
            result = process_file(filepath, print_func)
            results.append(result)
            if result['changes']:
                ok += 1

            if song_card_callback:
                song_card_callback({
                    'name': filepath.name,
                    'status': 'completed',
                    'artist': result.get('artist', ''),
                    'title': result.get('title', ''),
                    'album': result.get('album', ''),
                    'quality': result.get('quality', ''),
                    'changes': result['changes'],
                    'errors': result['errors'],
                    'duration': '',
                    'file_size': f"{filepath.stat().st_size / (1024*1024):.1f} MB"
                })

        except Exception as e:
            print_func(f"❌ Error processing {filepath.name}: {e}")
            results.append({
                'file': filepath, 'changes': [], 'errors': [str(e)],
                'artist': '', 'title': '', 'album': '', 'quality': ''
            })

            if song_card_callback:
                song_card_callback({
                    'name': filepath.name,
                    'status': 'error',
                    'artist': '', 'title': '', 'album': '',
                    'changes': [], 'errors': [str(e)],
                    'duration': '',
                    'file_size': f"{filepath.stat().st_size / (1024*1024):.1f}MB"
                })

        if progress_callback:
            progress_callback(i, len(files), str(filepath.name))

    print_func(f"\n\n{'═' * 50}")
    print_func(f"✅ Done — {ok}/{len(results)} files updated")

    # Console path: ask user; GUI handles its own report via the dialog
    if ask_report:
        print_change_summary(results)
        answer = input("\nGenerate Excel report? [y/N]: ").strip().lower()
        if answer == 'y':
            report_path = music_dir / f"orpheus_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            generate_excel_report(results, report_path)
        else:
            print("Report skipped.")

    return results

# ─── CONSOLE MENU ─────────────────────────────────────────────────────────────

def _parse_path(raw):
    """Normalise a user-supplied path: strip quotes, expand ~, resolve."""
    cleaned = raw.strip().strip('"').strip("'").strip()
    if not cleaned:
        return None
    return Path(cleaned).expanduser().resolve()

def _suggest_similar(path):
    """Return up to 3 sibling directories whose names are similar."""
    parent = path.parent
    name = path.name.lower()
    if not parent.exists():
        return []
    prefix = name[:max(1, len(name) // 2)]  # match at least half the name
    suggestions = [
        str(child)
        for child in sorted(parent.iterdir())
        if child.is_dir() and prefix in child.name.lower() and child != path
    ]
    return suggestions[:3]

def _count_music_files(folder):
    """Return the total number of supported music files found in folder."""
    count = 0
    for ext in EXTENSIONS:
        count += len(list(Path(folder).rglob(f"*{ext}")))
    return count

def console_menu():
    """Interactive console menu for running without the GUI."""
    music_dir = MUSIC_DIR  # may be empty string from config

    EXIT_WORDS = {'3', 'q', 'quit', 'exit', 'salir'}

    while True:
        print("\n" + "=" * 52)
        print("  🎵 Orpheus — Music Metadata Updater")
        print("=" * 52)
        folder_display = music_dir if music_dir else "(not set)"
        print(f"  Folder : {folder_display}")
        print(f"  Mode   : {'DRY-RUN (simulation)' if DRY_RUN else 'Live (files will be modified)'}")
        print()
        print("  [1] Set music folder")
        print("  [2] Start processing")
        print("  [3] Exit")
        print("=" * 52)

        raw_choice = input("  Select an option [1/2/3]: ").strip()
        choice = raw_choice.lower()

        # ── Exit shortcuts ────────────────────────────────────────
        if choice in EXIT_WORDS:
            print("  Goodbye.")
            sys.exit(0)

        # ── Option 1: set folder ──────────────────────────────────
        elif choice == '1':
            while True:
                raw = input("  Enter path to music folder (or 'back' to cancel): ")
                if raw.strip().lower() in ('back', 'b', 'cancel'):
                    break

                path = _parse_path(raw)
                if path is None:
                    print("  No path entered. Please type a folder path.")
                    continue

                if path.exists() and path.is_dir():
                    count = _count_music_files(path)
                    music_dir = str(path)
                    print(f"  Folder set to: {music_dir}")
                    if count == 0:
                        print(f"  Warning: no music files found in that folder.")
                        print(f"  Supported formats: {', '.join(sorted(EXTENSIONS))}")
                    else:
                        print(f"  {count} music file(s) found.")
                    break
                else:
                    print(f"  Folder not found: {path}")
                    suggestions = _suggest_similar(path)
                    if suggestions:
                        print("  Did you mean one of these?")
                        for s in suggestions:
                            print(f"    -> {s}")
                    else:
                        parent = path.parent
                        if parent.exists():
                            print(f"  Parent folder exists: {parent}")
                            print("  Check for typos in the folder name.")
                        else:
                            print("  The parent path also does not exist.")
                            print("  Please enter the full absolute path, e.g.:")
                            print("    /home/user/Music")

        # ── Option 2: start processing ────────────────────────────
        elif choice == '2':
            if not music_dir:
                print("  No folder selected.")
                print("  Use option [1] to set the music folder first.")
                continue

            folder_path = Path(music_dir)
            if not folder_path.exists():
                print(f"  The selected folder no longer exists: {music_dir}")
                print("  Use option [1] to set a valid folder.")
                music_dir = ''
                continue

            count = _count_music_files(music_dir)
            if count == 0:
                print(f"  No supported music files found in: {music_dir}")
                print(f"  Supported formats: {', '.join(sorted(EXTENSIONS))}")
                print("  Check the folder path or choose a different one with [1].")
                continue

            print()
            main(music_dir=music_dir, ask_report=True)

        # ── Unrecognised input ─────────────────────────────────────
        else:
            # Helpful hint if user typed a path instead of a number
            if os.sep in raw_choice or (len(raw_choice) > 2 and raw_choice[1] == ':'):
                print(f"  That looks like a path.")
                print("  Use option [1] to set the music folder.")
            elif len(raw_choice) > 1:
                print(f"  '{raw_choice}' is not a valid option.")
                print("  Please type 1, 2, or 3.")
            else:
                print("  Invalid option. Please enter 1, 2, or 3.")


if __name__ == "__main__":
    try:
        if len(sys.argv) > 1 and sys.argv[1] == '--gui':
            run_gui(main, generate_excel_report)
        else:
            console_menu()
    except (KeyboardInterrupt, SystemExit):
        pass
